# app.py
from flask import Flask, request, render_template, send_file, url_for, jsonify
from bk import run_inference
import numpy as np
import cv2
import os
import uuid
import shutil
import json
from threading import Thread, Lock
import copy
from datetime import datetime  # dùng để đặt tên file theo ngày giờ

from PIL import Image as PILImage  # dùng để kiểm tra/convert MPO -> JPG
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment  # dùng để style header & ô

app = Flask(__name__)

# Thư mục lưu kết quả theo session
SESSIONS_ROOT = os.path.join("check", "sessions")
os.makedirs(SESSIONS_ROOT, exist_ok=True)

# Các key public để show ra UI & xuất Excel
PUBLIC_KEYS = ("arms", "smile", "shoes", "legs", "scollar", "eyebrow_hair", "ear")

# Lưu trạng thái của từng session trong RAM
SESSION_STATUS = {}
SESSION_LOCK = Lock()


def _to_py_primitive(x):
    """
    Chuẩn hoá giá trị để JSON-safe:
    - numpy scalar -> python scalar
    - NaN -> None
    - np.bool_ -> bool
    - bool -> 'pass' / 'fail'
    """
    import math
    import numpy as np

    if x is None:
        return None
    if isinstance(x, np.generic):
        x = x.item()
    if isinstance(x, float) and math.isnan(x):
        return None
    if isinstance(x, (np.bool_,)):
        x = bool(x)
    if isinstance(x, bool):
        return "pass" if x else "fail"
    if isinstance(x, (int, float)):
        return x
    if isinstance(x, str):
        return x
    return str(x)


def process_session_async(session_id: str):
    """
    Thread nền: đọc orig_*.jpg trong session_dir, chạy run_inference,
    cập nhật SESSION_STATUS[session_id]["items"].
    """
    session_dir = os.path.join(SESSIONS_ROOT, session_id)

    with SESSION_LOCK:
        sess = SESSION_STATUS.get(session_id)
        if not sess:
            return
        gender = sess.get("gender", "male")
        items_snapshot = list(sess["items"])

    for item in items_snapshot:
        idx = item["index"]

        with SESSION_LOCK:
            sess = SESSION_STATUS.get(session_id)
            if not sess:
                return
            cur_item = sess["items"][idx]
            # nếu không còn pending thì bỏ qua
            if cur_item["status"] != "pending":
                continue
            cur_item["status"] = "processing"

        orig_path = os.path.join(session_dir, f"orig_{idx}.jpg")
        if not os.path.exists(orig_path):
            with SESSION_LOCK:
                cur_item["status"] = "error"
                cur_item["error_message"] = "Không tìm thấy file gốc."
            continue

        img = cv2.imread(orig_path)
        if img is None:
            with SESSION_LOCK:
                cur_item["status"] = "error"
                cur_item["error_message"] = "Ảnh không hợp lệ hoặc bị lỗi."
            continue

        # Gọi pipeline chính
        try:
            output_path, results = run_inference(img, test_image_path=None, gender=gender)
        except Exception as e:
            with SESSION_LOCK:
                cur_item["status"] = "error"
                cur_item["error_message"] = f"Lỗi xử lý ảnh: {e}"
            continue

        # Lỗi chung từ pipeline (vd: không thấy pose, v.v.)
        if results.get("_error"):
            with SESSION_LOCK:
                cur_item["status"] = "error"
                cur_item["error_message"] = results["_error"]
            continue

        # Tổng hợp shoes nếu chưa có
        if "shoes" not in results:
            left_shoe = results.get("left_shoe", "missing")
            right_shoe = results.get("right_shoe", "missing")
            results["shoes"] = "pass" if (left_shoe == "pass" and right_shoe == "pass") else "fail"

        # Lấy subset public
        results_public = {k: results.get(k) for k in PUBLIC_KEYS}
        for k, v in list(results_public.items()):
            results_public[k] = _to_py_primitive(v)

        # Lưu ảnh annotate vào session_dir: img_{idx}.jpg
        dest_path = os.path.join(session_dir, f"img_{idx}.jpg")
        if output_path and os.path.exists(output_path):
            try:
                shutil.copy2(output_path, dest_path)
            except Exception:
                shutil.copy(output_path, dest_path)
        else:
            # Fallback: lưu ảnh gốc nếu không có file annotate
            cv2.imwrite(dest_path, img)

        with SESSION_LOCK:
            cur_item["status"] = "ok"
            cur_item["results"] = results_public
            cur_item["error_message"] = None

    # Đánh dấu session done
    with SESSION_LOCK:
        sess = SESSION_STATUS.get(session_id)
        if sess:
            sess["done"] = True


@app.route('/')
def index():
    return render_template("index.html")


@app.route('/upload', methods=['POST'])
def upload():
    """
    - Nhận 1 ảnh đơn (image) hoặc folder nhiều ảnh (images).
    - Lưu file gốc vào session_dir, tạo list items với status='pending'.
    - Trả ngay result.html.
    - Thread nền process_session_async sẽ xử lý từng ảnh.
    """
    # Ưu tiên folder / nhiều ảnh
    files_multi = request.files.getlist("images")
    file_single = request.files.get("image")

    selected_files = []
    if files_multi and any(f.filename for f in files_multi):
        selected_files = [f for f in files_multi if f and f.filename]
    elif file_single and file_single.filename:
        selected_files = [file_single]

    if not selected_files:
        return render_template(
            "error.html",
            message="❌ Không có file nào được tải lên.",
            back_url=url_for('index')
        ), 400

    # Giới tính
    gender = request.form.get("gender", "male").lower()
    if gender not in ("male", "female"):
        gender = "male"

    # Tạo session id riêng cho đợt upload này
    session_id = uuid.uuid4().hex
    session_dir = os.path.join(SESSIONS_ROOT, session_id)
    os.makedirs(session_dir, exist_ok=True)

    items = []

    for idx, file in enumerate(selected_files):
        filename = os.path.basename(file.filename or f"image_{idx+1}.jpg")
        item = {
            "index": idx,
            "filename": filename,
            "status": "pending",        # 'pending' | 'processing' | 'ok' | 'error'
            "results": None,
            "error_message": None,
        }

        # Lưu file gốc để xử lý sau (luôn lưu thành orig_{idx}.jpg)
        orig_path = os.path.join(session_dir, f"orig_{idx}.jpg")
        try:
            file.save(orig_path)
        except Exception as e:
            item["status"] = "error"
            item["error_message"] = f"Lỗi lưu file: {e}"

        items.append(item)

    # Nếu tất cả đều lỗi lưu file
    if not any(it["status"] == "pending" for it in items):
        msg = items[0]["error_message"] if items and items[0]["error_message"] else "Không xử lý được ảnh nào."
        return render_template(
            "error.html",
            message=msg,
            back_url=url_for('index')
        ), 422

    # Gán display_index để dùng trên frontend
    for i, it in enumerate(items):
        it["display_index"] = i

    initial_index = 0
    initial_item = items[initial_index]

    # Lưu trạng thái session vào RAM
    with SESSION_LOCK:
        SESSION_STATUS[session_id] = {
            "items": items,
            "done": False,
            "gender": gender,
        }

    # Khởi chạy thread nền
    t = Thread(target=process_session_async, args=(session_id,), daemon=True)
    t.start()

    # Render result.html ngay lập tức
    return render_template(
        "result.html",
        session_id=session_id,
        items=items,
        initial_index=initial_index,
        initial_item=initial_item,
    )


@app.route('/result/<session_id>/<int:idx>')
def result_image(session_id, idx: int):
    """
    Trả ảnh annotate cho từng ảnh trong 1 session.
    URL mẫu: /result/<session_id>/0, /result/<session_id>/1, ...
    """
    session_dir = os.path.join(SESSIONS_ROOT, session_id)
    path = os.path.join(session_dir, f"img_{idx}.jpg")
    if not os.path.exists(path):
        return "❌ Không tìm thấy ảnh kết quả.", 404
    return send_file(path, mimetype='image/jpeg')


@app.route('/session/<session_id>/excel')
def session_excel(session_id):
    """Tạo file Excel tổng hợp kết quả cho 1 session.
    - Mỗi hàng tương ứng 1 ảnh đã upload.
    - Cột: STT, File name, các tiêu chí trong PUBLIC_KEYS, và cột cuối là ảnh gốc.
    - Chỉ cho phép tải khi session đã xử lý xong (sess["done"] == True).
    """
    # Lấy snapshot trong lock
    with SESSION_LOCK:
        sess = SESSION_STATUS.get(session_id)
        if not sess:
            return "❌ Session không tồn tại.", 404

        if not sess.get("done", False):
            # 409: Conflict / chưa sẵn sàng
            return "⏳ Session vẫn đang xử lý, chưa thể xuất Excel.", 409

        items_snapshot = copy.deepcopy(sess["items"])

    session_dir = os.path.join(SESSIONS_ROOT, session_id)

    # Khởi tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Header: STT + File name + keys + Image
    headers = ["STT", "File name"] + list(PUBLIC_KEYS) + ["Image"]
    ws.append(headers)

    # Style header: xanh dương nhạt, in đậm, chữ to, căn giữa
    header_fill = PatternFill(start_color="B7D8FF", end_color="B7D8FF", fill_type="solid")
    header_font = Font(bold=True, size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Đặt độ rộng cột cơ bản
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        if header == "Image":
            ws.column_dimensions[col_letter].width = 30
        elif header == "File name":
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = 15

    # Style cho ô PASS / SMILE vs FAIL / NO_SMILE
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # xanh lá nhạt
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # đỏ nhạt
    center_align = Alignment(horizontal="center", vertical="center")

    # Ghi từng dòng
    for row_idx, item in enumerate(items_snapshot, start=1):
        filename = item.get("filename") or f"image_{row_idx}.jpg"
        results = item.get("results") or {}
        status = item.get("status", "")

        # STT là row_idx
        row_vals = [row_idx, filename]

        # Nếu ảnh xử lý OK -> fill các tiêu chí
        if status == "ok" and results:
            for key in PUBLIC_KEYS:
                val = results.get(key)
                row_vals.append("" if val is None else str(val))
        else:
            # Nếu lỗi/pending -> để trống các cột tiêu chí
            for _ in PUBLIC_KEYS:
                row_vals.append("")

        # Placeholder cho cột Image
        row_vals.append("")
        ws.append(row_vals)
        excel_row = ws.max_row

        # Căn giữa cột STT
        ws.cell(row=excel_row, column=1).alignment = center_align

        # Tô màu cho các cột kết quả (từ cột 3 tới 2 + len(PUBLIC_KEYS))
        start_col = 3
        end_col = 2 + len(PUBLIC_KEYS)
        for col_idx in range(start_col, end_col + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            value_str = (str(cell.value).strip().lower()
                         if cell.value is not None else "")
            if value_str in ("pass", "smile"):
                cell.fill = pass_fill
            elif value_str in ("fail", "no_smile"):
                cell.fill = fail_fill
            # căn giữa các ô kết quả
            cell.alignment = center_align

        # Thử chèn ảnh gốc vào cột cuối
        try:
            orig_index = item.get("index", row_idx - 1)
            orig_path = os.path.join(session_dir, f"orig_{orig_index}.jpg")

            if os.path.exists(orig_path):
                img_path_for_excel = orig_path

                # Mở bằng Pillow để kiểm tra định dạng (MPO -> convert sang JPG)
                try:
                    with PILImage.open(orig_path) as im:
                        fmt = (im.format or "").upper()
                        if fmt == "MPO":
                            # convert MPO -> JPG tạm cho Excel
                            tmp_path = os.path.join(session_dir, f"excel_img_{orig_index}.jpg")
                            im.convert("RGB").save(tmp_path, format="JPEG")
                            img_path_for_excel = tmp_path
                except Exception as e:
                    # Nếu không mở được bằng PIL, vẫn dùng orig_path (JPG thường)
                    print(f"Lỗi đọc ảnh bằng PIL ({orig_path}):", e)

                img = XLImage(img_path_for_excel)

                # Thu nhỏ chiều cao ảnh để bảng không quá cao
                max_height = 160
                try:
                    if img.height and img.height > max_height:
                        scale = max_height / float(img.height)
                        img.height = max_height
                        img.width = int(img.width * scale)
                except Exception:
                    # Nếu không đọc được kích thước thì bỏ qua resize
                    pass

                img_col = get_column_letter(len(headers))
                ws.add_image(img, f"{img_col}{excel_row}")
                # tăng chiều cao dòng cho vừa ảnh
                ws.row_dimensions[excel_row].height = 120
        except Exception as e:
            # Nếu chèn ảnh lỗi thì bỏ qua, không chặn việc xuất Excel
            print("Lỗi chèn ảnh vào Excel:", e)
            continue

    # Lưu file Excel vào thư mục session với tên theo ngày + giờ
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"uniform_results_{timestamp}.xlsx"
    excel_path = os.path.join(session_dir, excel_filename)
    wb.save(excel_path)

    return send_file(
        excel_path,
        as_attachment=True,
        download_name=excel_filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route('/session/<session_id>/status')
def session_status(session_id):
    """
    API cho front-end poll:
    trả về {"items": [...], "done": true/false}
    """
    with SESSION_LOCK:
        sess = SESSION_STATUS.get(session_id)
        if not sess:
            return jsonify({"error": "session_not_found"}), 404

        data = {
            "items": copy.deepcopy(sess["items"]),
            "done": bool(sess.get("done", False)),
        }

    return jsonify(data)


if __name__ == '__main__':
    app.run(debug=True, port=5001)
